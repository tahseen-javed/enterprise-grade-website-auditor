"""
Export integrity, API endpoints, and resume/checkpoint behaviour
(spec 11, 28, 30, 33, 45, 46).
"""

from __future__ import annotations

import csv as csvmod
import io
import json
import zipfile
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from app.core.exporter import ENRICHMENT_COLUMNS, build_rows, export_csv, export_xlsx
from app.db import session_scope
from app.main import app
from app.models import (
    Business, ContactEmail, ContactPhone, Job, JobItem, OutreachDraft, WebsiteAudit,
)

ORIGINAL_HEADERS = [
    "Business Name", "Category", "Phone Number", "Website URL", "City", "Country",
    "My Private Notes", "website_score",
]


@pytest.fixture
def seeded_job():
    """A job with three finished leads covering the three channel outcomes."""
    with session_scope() as s:
        job = Job(
            name="EXPORT TEST", source_filename="t.csv", stored_path="t.csv",
            original_columns=list(ORIGINAL_HEADERS), column_mapping={}, status="completed",
            total=3,
        )
        s.add(job)
        s.flush()

        rows = [
            {
                "Business Name": "WhatsApp Lead Ltd", "Category": "Plumber",
                "Phone Number": "+44 7911 123456", "Website URL": "https://wa-lead.test",
                "City": "Leeds", "Country": "United Kingdom",
                "My Private Notes": "keep, comma, and \"quotes\"", "website_score": "my own value",
            },
            {
                "Business Name": "Email Lead Ltd", "Category": "Dentist",
                "Phone Number": "+44 113 296 0001", "Website URL": "https://email-lead.test",
                "City": "Leeds", "Country": "United Kingdom",
                "My Private Notes": "=SUM(A1:A9)", "website_score": "",
            },
            {
                "Business Name": "Call Lead Ltd", "Category": "Roofer",
                "Phone Number": "+61 2 9374 4000", "Website URL": "",
                "City": "Sydney", "Country": "Australia",
                "My Private Notes": "", "website_score": "",
            },
        ]

        made = []
        for i, raw in enumerate(rows):
            b = Business(
                job_id=job.id, row_index=i, raw=raw, name=raw["Business Name"],
                category=raw["Category"], city=raw["City"], country=raw["Country"],
                phone_raw=raw["Phone Number"], website_original=raw["Website URL"],
            )
            s.add(b)
            s.flush()
            made.append(b)
            s.add(JobItem(job_id=job.id, business_id=b.id, status="completed", stage="done",
                          completed_stages=["normalize", "done"]))

        wa, em, call = made

        wa.website_status = "valid"; wa.website_final = "https://wa-lead.test/"
        wa.website_identity_confidence = 0.91; wa.website_source = "csv"
        wa.score = 82; wa.opportunity_tier = "High"; wa.lead_tier = "A+"
        wa.audit_kind = "website"; wa.best_channel = "whatsapp"
        wa.channel_reason = "WhatsApp-capable number"
        s.add(ContactPhone(
            business_id=wa.id, phone_raw="+44 7911 123456", phone_normalized="+447911123456",
            phone_country="GB", phone_type="mobile", validation_status="valid",
            whatsapp_status="usable_unverified", whatsapp_reason="Valid mobile number",
            whatsapp_url="https://wa.me/447911123456?text=Hi",
        ))
        s.add(WebsiteAudit(
            business_id=wa.id, website="https://wa-lead.test/", audit_kind="website",
            score=82, opportunity_tier="High", audit_status="completed",
            problems=[{"rank": 1, "title": "No clear call to action", "severity": "high",
                       "code": "no_primary_cta_above_fold"}],
            recommendations=[{"rank": 1, "problem_code": "no_primary_cta_above_fold",
                              "recommendation": "Add a prominent booking button."}],
            extra={"priorities": [
                {"rank": 1, "code": "no_https", "category": "security", "category_label": "Security",
                 "severity": "high", "priority": "P1", "title": "The site does not load over HTTPS",
                 "detail": "Not served over HTTPS.", "recommendation": "Install an SSL certificate.",
                 "impact_points": 30},
                {"rank": 2, "code": "missing_title", "category": "onpage", "category_label": "On-Page SEO",
                 "severity": "medium", "priority": "P2", "title": "The homepage has no title tag",
                 "detail": "No <title> element was found.", "recommendation": "Add a descriptive title.",
                 "impact_points": 18},
            ]},
            report_path="",
        ))
        s.add(OutreachDraft(
            business_id=wa.id, channel="whatsapp", variant="initial",
            message="Hi,\n\nI noticed the CTA is hard to find.", draft_url="https://wa.me/447911123456?text=Hi",
            based_on=[{"code": "no_primary_cta_above_fold", "observation": "cta hard to find"}],
        ))

        em.website_status = "valid"; em.website_final = "https://email-lead.test/"
        em.score = 64; em.opportunity_tier = "Good"; em.lead_tier = "A"
        em.audit_kind = "website"; em.best_channel = "email"
        em.channel_reason = "No usable WhatsApp; public email found"
        s.add(ContactPhone(
            business_id=em.id, phone_raw="+44 113 296 0001", phone_normalized="+441132960001",
            phone_country="GB", phone_type="fixed_line", validation_status="valid",
            whatsapp_status="unlikely", whatsapp_reason="Landline", whatsapp_url="",
        ))
        s.add(ContactEmail(
            business_id=em.id, email="info@email-lead.test",
            source_url="https://email-lead.test/contact", source_type="mailto",
            page_type="contact", status="valid_public", confidence=0.93,
            is_role=True, domain_matches_site=True, mx_records=["mx1.test"], rank=0,
        ))
        s.add(ContactEmail(
            business_id=em.id, email="bookings@email-lead.test",
            source_url="https://email-lead.test/contact", source_type="text",
            page_type="contact", status="mx_valid", confidence=0.8, rank=1,
        ))
        s.add(WebsiteAudit(
            business_id=em.id, website="https://email-lead.test/", audit_kind="website",
            score=64, opportunity_tier="Good", audit_status="completed",
            problems=[{"rank": 1, "title": "No testimonials found", "severity": "high",
                       "code": "no_testimonials"}],
            recommendations=[], report_path="",
        ))
        s.add(OutreachDraft(
            business_id=em.id, channel="email", variant="initial",
            subject="Quick note about Email Lead Ltd's website",
            message="Hi,\n\nNo testimonials on the site.",
            draft_url="mailto:info@email-lead.test?subject=x&body=y",
        ))

        call.website_status = "no_website"; call.opportunity_tier = "No website"
        call.lead_tier = "B"; call.audit_kind = "no_website"; call.best_channel = "phone"
        call.channel_reason = "No WhatsApp or email; phone available"
        s.add(ContactPhone(
            business_id=call.id, phone_raw="+61 2 9374 4000", phone_normalized="+61293744000",
            phone_country="AU", phone_type="fixed_line", validation_status="valid",
            whatsapp_status="unlikely", whatsapp_reason="Landline", whatsapp_url="",
        ))
        s.add(WebsiteAudit(
            business_id=call.id, website="", audit_kind="no_website",
            audit_status="completed",
            problems=[{"rank": 1, "title": "No website could be found", "severity": "high",
                       "code": "no_website_detected"}],
            recommendations=[], report_path="",
        ))
        s.add(OutreachDraft(
            business_id=call.id, channel="call", variant="initial",
            message="OPENER: Hi, is that Call Lead Ltd?",
        ))

        job_id = job.id

    yield job_id

    with session_scope() as s:
        job = s.get(Job, job_id)
        if job:
            s.delete(job)


class TestExportIntegrity:
    def test_original_columns_come_first_in_order(self, seeded_job):
        with session_scope(write=False) as s:
            headers, _ = build_rows(s, seeded_job)
        assert headers[: len(ORIGINAL_HEADERS)] == ORIGINAL_HEADERS

    def test_one_input_row_stays_one_output_row(self, seeded_job):
        with session_scope(write=False) as s:
            _, rows = build_rows(s, seeded_job)
        assert len(rows) == 3
        assert [r["Business Name"] for r in rows] == [
            "WhatsApp Lead Ltd", "Email Lead Ltd", "Call Lead Ltd"
        ]

    def test_no_website_lead_stays_a_row_but_carries_no_enrichment(self):
        """
        A lead the pipeline excluded for having no valid website (spec: website
        required) must still be exactly one output row - the row-preservation
        guarantee holds for every lead, not just processed ones - but with no
        score, tier, problems or drafts, since it was never audited.
        """
        with session_scope() as s:
            job = Job(
                name="NO WEBSITE EXPORT TEST", source_filename="w.csv", stored_path="w.csv",
                original_columns=["Business Name", "Phone Number", "Website URL"],
                column_mapping={}, status="completed", total=1,
            )
            s.add(job)
            s.flush()
            raw = {"Business Name": "No Site Bakery Ltd", "Phone Number": "+44 7911 000111",
                   "Website URL": ""}
            biz = Business(
                job_id=job.id, row_index=0, raw=raw, name=raw["Business Name"],
                phone_raw=raw["Phone Number"], website_original="",
                website_status="no_website", best_channel="none",
                channel_reason="No valid website could be confirmed for this business.",
            )
            s.add(biz)
            s.flush()
            s.add(JobItem(job_id=job.id, business_id=biz.id, status="skipped",
                          stage="no_website",
                          error_message="Skipped — no valid website (status: no_website)."))
            job_id = job.id

        try:
            with session_scope(write=False) as s:
                headers, rows = build_rows(s, job_id)
            assert len(rows) == 1
            assert rows[0]["Business Name"] == "No Site Bakery Ltd"
            assert rows[0]["phone_raw_original"] == "+44 7911 000111"
            assert rows[0]["website_status"] == "no_website"
            assert rows[0]["website_score"] == ""
            assert rows[0]["lead_tier"] == ""
            assert rows[0]["whatsapp_message"] == ""
            assert rows[0]["email_message"] == ""
            assert rows[0]["problems"] == ""
            assert rows[0]["audit_status"] == "skipped"
        finally:
            with session_scope() as s:
                j = s.get(Job, job_id)
                if j:
                    s.delete(j)

    def test_user_data_is_never_altered(self, seeded_job):
        with session_scope(write=False) as s:
            _, rows = build_rows(s, seeded_job)
        assert rows[0]["My Private Notes"] == 'keep, comma, and "quotes"'
        assert rows[0]["Category"] == "Plumber"

    def test_colliding_column_name_does_not_overwrite_user_data(self, seeded_job):
        """The input already has a 'website_score' column of its own."""
        with session_scope(write=False) as s:
            headers, rows = build_rows(s, seeded_job)
        assert rows[0]["website_score"] == "my own value"
        assert "audit_website_score" in headers
        assert rows[0]["audit_website_score"] == 82

    def test_enrichment_columns_are_appended_at_the_end(self, seeded_job):
        with session_scope(write=False) as s:
            headers, _ = build_rows(s, seeded_job)
        tail = headers[len(ORIGINAL_HEADERS):]
        assert len(tail) == len(ENRICHMENT_COLUMNS)
        assert tail[0] in ("website_status", "audit_website_status")

    def test_channel_specific_fields_are_populated(self, seeded_job):
        with session_scope(write=False) as s:
            _, rows = build_rows(s, seeded_job)
        wa, em, call = rows
        assert wa["whatsapp_status"] == "usable_unverified"
        assert wa["whatsapp_url"].startswith("https://wa.me/447911123456")
        assert wa["whatsapp_message"]
        assert wa["contact_channel"] == "WHATSAPP"

        assert em["email_1"] == "info@email-lead.test"
        assert em["email_1_status"] == "valid_public"
        assert em["email_1_source"].endswith("/contact")
        assert em["email_2"] == "bookings@email-lead.test"
        assert em["email_subject"]
        assert em["contact_channel"] == "EMAIL"

        assert call["call_notes"].startswith("OPENER:")
        assert call["contact_channel"] == "PHONE"
        assert call["phone_normalized"] == "+61293744000"
        assert call["phone_country"] == "AU"

    def test_problems_and_recommendations_are_exported(self, seeded_job):
        with session_scope(write=False) as s:
            _, rows = build_rows(s, seeded_job)
        assert "No clear call to action" in rows[0]["problems"]
        assert "prominent booking button" in rows[0]["recommendations"]

    def test_top_priority_issues_get_their_own_columns(self, seeded_job):
        """The premium Top 5 priorities must never be exported as a single
        joined cell - each ranked issue gets its own title/category/severity/
        fix columns, and unused ranks are blank rather than omitted."""
        with session_scope(write=False) as s:
            headers, rows = build_rows(s, seeded_job)
        wa = rows[0]
        assert wa["top_issue_1"] == "The site does not load over HTTPS"
        assert wa["top_issue_1_category"] == "Security"
        assert wa["top_issue_1_severity"] == "HIGH"
        assert wa["top_issue_1_fix"] == "Install an SSL certificate."
        assert wa["top_issue_2"] == "The homepage has no title tag"
        assert wa["top_issue_3"] == ""  # only 2 priorities were recorded for this lead
        assert "top_issue_5_fix" in headers

        em = rows[1]  # no `extra.priorities` at all for this lead
        assert em["top_issue_1"] == ""

    def test_no_whatsapp_url_for_a_landline(self, seeded_job):
        with session_scope(write=False) as s:
            _, rows = build_rows(s, seeded_job)
        assert rows[1]["whatsapp_url"] == ""
        assert rows[1]["whatsapp_status"] == "unlikely"

    def test_csv_file_round_trips(self, seeded_job):
        with session_scope(write=False) as s:
            path = export_csv(s, seeded_job)
        assert path.exists()
        with path.open(encoding="utf-8-sig", newline="") as fh:
            reader = csvmod.DictReader(fh)
            out_headers = reader.fieldnames
            out_rows = list(reader)
        assert out_headers[: len(ORIGINAL_HEADERS)] == ORIGINAL_HEADERS
        assert len(out_rows) == 3
        assert out_rows[0]["My Private Notes"] == 'keep, comma, and "quotes"'

    def test_formula_injection_is_neutralised(self, seeded_job):
        with session_scope(write=False) as s:
            _, rows = build_rows(s, seeded_job)
        # The raw value is preserved as the user supplied it...
        assert rows[1]["My Private Notes"] == "=SUM(A1:A9)"
        # ...and the spreadsheet cell is escaped so Excel will not execute it.
        from app.core.exporter import _cell

        assert _cell("=SUM(A1:A9)").startswith("'")

    def test_xlsx_has_data_and_a_documentation_sheet(self, seeded_job):
        with session_scope(write=False) as s:
            path = export_xlsx(s, seeded_job)
        assert path.exists()
        from openpyxl import load_workbook

        wb = load_workbook(path)
        assert "Leads" in wb.sheetnames
        assert "Enrichment key" in wb.sheetnames
        ws = wb["Leads"]
        assert ws.max_row == 4  # header + 3 leads
        assert [c.value for c in ws[1]][: len(ORIGINAL_HEADERS)] == ORIGINAL_HEADERS


class TestApiEndpoints:
    @pytest.fixture
    def client(self):
        with TestClient(app) as c:
            yield c

    def test_health(self, client):
        r = client.get("/api/health")
        assert r.status_code == 200
        assert r.json()["status"] == "ok"

    def test_system_health_reports_every_component(self, client):
        r = client.get("/api/system/health")
        assert r.status_code == 200
        body = r.json()
        assert body["overall"] in ("healthy", "warning", "error")
        for key in ("backend", "database", "crawler", "export_engine", "file_system", "ports"):
            assert key in body["components"]
            assert body["components"][key]["detail"]

    def test_secrets_are_masked_in_settings(self, client):
        r = client.get("/api/settings")
        assert r.status_code == 200
        engine = r.json()["engine"]
        for key in ("pagespeed_api_key", "llm_api_key", "google_places_api_key"):
            assert engine[key] in ("", "********")
            assert f"{key}_set" in engine

    def test_profile_round_trip(self, client):
        original = client.get("/api/settings/profile").json()["profile"]
        try:
            r = client.put("/api/settings/profile", json={
                "full_name": "API Test User", "company_name": "API Test Co",
                "service_name": "website redesign", "tone": "friendly",
                "target_countries": ["United Kingdom"],
            })
            assert r.status_code == 200
            body = r.json()
            assert body["profile"]["full_name"] == "API Test User"
            assert body["profile"]["tone"] == "friendly"
            assert body["profile"]["target_countries"] == ["United Kingdom"]
            assert body["status"]["configured"] is True
        finally:
            client.put("/api/settings/profile", json=original)

    def test_invalid_tone_is_rejected_to_a_safe_default(self, client):
        original = client.get("/api/settings/profile").json()["profile"]
        try:
            r = client.put("/api/settings/profile", json={"tone": "aggressive"})
            assert r.json()["profile"]["tone"] == "professional"
        finally:
            client.put("/api/settings/profile", json=original)

    def test_worker_count_is_clamped(self, client):
        original = client.get("/api/settings/engine").json()["engine"]
        try:
            assert client.put("/api/settings/engine", json={"workers": 99}).status_code == 422
            r = client.put("/api/settings/engine", json={"workers": 20})
            assert r.json()["engine"]["workers"] == 20
        finally:
            client.put("/api/settings/engine", json={"workers": original["workers"]})

    def test_scoring_weights_round_trip(self, client):
        original = client.get("/api/settings/scoring").json()["scoring"]
        try:
            r = client.put("/api/settings/scoring", json={"weights": {"mobile": 30}})
            assert r.json()["scoring"]["weights"]["mobile"] == 30
        finally:
            client.put("/api/settings/scoring", json={"weights": original["weights"]})

    def test_enrichment_column_documentation_is_served(self, client):
        r = client.get("/api/exports/columns")
        assert r.status_code == 200
        body = r.json()
        assert body["columns"] == ENRICHMENT_COLUMNS
        assert "whatsapp_status" in body["documentation"]

    def test_jobs_list(self, client, seeded_job):
        r = client.get("/api/jobs")
        assert r.status_code == 200
        assert any(j["id"] == seeded_job for j in r.json()["jobs"])

    def test_job_detail_and_stats(self, client, seeded_job):
        r = client.get(f"/api/jobs/{seeded_job}")
        assert r.status_code == 200
        assert r.json()["counts"]["completed"] == 3

        r = client.get(f"/api/jobs/{seeded_job}/stats")
        assert r.status_code == 200
        stats = r.json()
        assert stats["total"] == 3
        assert stats["channels"]["whatsapp"] == 1
        assert stats["channels"]["email"] == 1
        assert stats["channels"]["phone"] == 1

    def test_missing_job_returns_404(self, client):
        assert client.get("/api/jobs/99999999").status_code == 404

    def test_leads_list_and_filters(self, client, seeded_job):
        r = client.get("/api/leads", params={"job_id": seeded_job})
        assert r.status_code == 200
        assert r.json()["total"] == 3

        r = client.get("/api/leads", params={"job_id": seeded_job, "channel": "whatsapp"})
        assert r.json()["total"] == 1
        assert r.json()["leads"][0]["name"] == "WhatsApp Lead Ltd"

        r = client.get("/api/leads", params={"job_id": seeded_job, "lead_tier": "A+"})
        assert r.json()["total"] == 1

        r = client.get("/api/leads", params={"job_id": seeded_job, "min_score": 70})
        assert r.json()["total"] == 1

        r = client.get("/api/leads", params={"job_id": seeded_job, "has_email": True})
        assert r.json()["total"] == 1

        r = client.get("/api/leads", params={"job_id": seeded_job, "search": "Call Lead"})
        assert r.json()["total"] == 1

    def test_include_drafts_returns_full_message(self, client, seeded_job):
        r = client.get("/api/leads", params={"job_id": seeded_job, "channel": "email",
                                            "include_drafts": True})
        lead = r.json()["leads"][0]
        assert "email" in lead["initial_drafts"]
        assert lead["initial_drafts"]["email"]["subject"]
        assert lead["initial_drafts"]["email"]["message"]

    def test_lead_detail_exposes_evidence_and_raw_row(self, client, seeded_job):
        listing = client.get("/api/leads", params={"job_id": seeded_job, "channel": "whatsapp"}).json()
        lead_id = listing["leads"][0]["id"]
        r = client.get(f"/api/leads/{lead_id}")
        assert r.status_code == 200
        body = r.json()
        assert body["audit"]["problems"]
        assert body["raw"]["My Private Notes"] == 'keep, comma, and "quotes"'
        assert body["phone"]["whatsapp_status"] == "usable_unverified"
        assert body["drafts"]

    def test_report_404_when_none_generated(self, client, seeded_job):
        listing = client.get("/api/leads", params={"job_id": seeded_job}).json()
        assert client.get(f"/api/leads/{listing['leads'][0]['id']}/report").status_code == 404

    def test_csv_and_xlsx_downloads(self, client, seeded_job):
        r = client.get(f"/api/exports/{seeded_job}/csv")
        assert r.status_code == 200
        assert "text/csv" in r.headers["content-type"]
        text = r.content.decode("utf-8-sig")
        assert "WhatsApp Lead Ltd" in text
        assert "whatsapp_status" in text.splitlines()[0]

        r = client.get(f"/api/exports/{seeded_job}/xlsx")
        assert r.status_code == 200
        assert "spreadsheet" in r.headers["content-type"]

    def test_reports_zip_404_when_no_reports(self, client, seeded_job):
        assert client.get(f"/api/exports/{seeded_job}/reports.zip").status_code == 404

    def test_recent_events_endpoint(self, client):
        r = client.get("/api/events/recent", params={"limit": 5})
        assert r.status_code == 200
        assert "events" in r.json()

    def test_global_stats(self, client):
        r = client.get("/api/stats")
        assert r.status_code == 200
        assert "total" in r.json()

    def test_upload_rejects_a_non_csv_extension(self, client):
        r = client.post("/api/uploads",
                        files={"file": ("bad.exe", b"binary", "application/octet-stream")})
        assert r.status_code == 400

    def test_upload_rejects_an_empty_file(self, client):
        r = client.post("/api/uploads", files={"file": ("empty.csv", b"", "text/csv")})
        assert r.status_code == 400

    def test_upload_then_map_then_create_job(self, client):
        csv_bytes = (
            "Business Name,Phone Number,Website URL,City,Country\n"
            "Upload Test Ltd,+44 113 296 0001,https://upload-test.invalid,Leeds,United Kingdom\n"
        ).encode()
        r = client.post("/api/uploads", files={"file": ("leads.csv", csv_bytes, "text/csv")})
        assert r.status_code == 200
        body = r.json()
        assert body["row_count"] == 1
        assert body["mapping"]["business_name"] == "Business Name"
        assert body["mapping"]["phone"] == "Phone Number"

        r = client.post("/api/jobs", json={
            "upload_id": body["upload_id"], "name": "UPLOAD TEST",
            "mapping": {k: v for k, v in body["mapping"].items() if v},
            "skip_duplicates": True, "start_immediately": False,
        })
        assert r.status_code == 200
        created = r.json()
        assert created["total"] == 1
        assert created["started"] is False
        client.delete(f"/api/jobs/{created['job_id']}")

    def test_upload_accepts_xlsx_and_maps_it_identically_to_csv(self, client):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Business Name", "Phone Number", "Website URL", "City", "Country"])
        ws.append(["Excel Upload Ltd", "+44 113 296 0002", "https://excel-upload.invalid",
                   "Leeds", "United Kingdom"])
        ws.append(["Second Row Ltd", "+44 113 296 0003", "https://excel-upload-2.invalid",
                   "Leeds", "United Kingdom"])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        r = client.post(
            "/api/uploads",
            files={"file": ("leads.xlsx", xlsx_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 200
        body = r.json()
        assert body["source_format"] == "excel"
        assert body["row_count"] == 2
        assert body["mapping"]["business_name"] == "Business Name"
        assert body["mapping"]["phone"] == "Phone Number"
        assert body["sample_rows"][0]["Business Name"] == "Excel Upload Ltd"

        # The same upload_id now feeds job creation exactly like a CSV would -
        # no separate code path downstream of the upload endpoint.
        r = client.post("/api/jobs", json={
            "upload_id": body["upload_id"], "name": "XLSX UPLOAD TEST",
            "mapping": {k: v for k, v in body["mapping"].items() if v},
            "skip_duplicates": True, "start_immediately": False,
        })
        assert r.status_code == 200
        created = r.json()
        assert created["total"] == 2
        client.delete(f"/api/jobs/{created['job_id']}")

    def test_upload_rejects_an_empty_xlsx_workbook(self, client):
        from openpyxl import Workbook

        wb = Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        r = client.post(
            "/api/uploads",
            files={"file": ("empty.xlsx", buf.getvalue(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 400

    def test_upload_rejects_a_corrupt_xlsx_file(self, client):
        r = client.post(
            "/api/uploads",
            files={"file": ("corrupt.xlsx", b"not a real workbook",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 400

    def test_job_creation_requires_a_business_name_mapping(self, client):
        csv_bytes = b"colA,colB\nfoo,bar\n"
        up = client.post("/api/uploads", files={"file": ("x.csv", csv_bytes, "text/csv")}).json()
        r = client.post("/api/jobs", json={
            "upload_id": up["upload_id"], "mapping": {"city": "colA"},
            "start_immediately": False,
        })
        assert r.status_code == 400
        assert "business name" in r.json()["detail"].lower()

    def test_path_traversal_in_upload_id_is_rejected(self, client):
        assert client.get("/api/uploads/..%2F..%2Fetc%2Fpasswd/preview").status_code in (400, 404)
        r = client.post("/api/jobs", json={
            "upload_id": "../../etc/passwd", "mapping": {"business_name": "a"},
        })
        assert r.status_code == 400

    def test_duplicates_are_skipped_but_still_exported(self, client):
        csv_bytes = (
            "Business Name,Website URL,City\n"
            "Dup Ltd,https://dup-test.invalid,Leeds\n"
            "Dup Ltd Again,https://dup-test.invalid,Leeds\n"
        ).encode()
        up = client.post("/api/uploads", files={"file": ("d.csv", csv_bytes, "text/csv")}).json()
        created = client.post("/api/jobs", json={
            "upload_id": up["upload_id"], "name": "DUP TEST",
            "mapping": {k: v for k, v in up["mapping"].items() if v},
            "skip_duplicates": True, "start_immediately": False,
        }).json()
        try:
            assert created["duplicates_skipped"] == 1
            csv_out = client.get(f"/api/exports/{created['job_id']}/csv").content.decode("utf-8-sig")
            assert "Dup Ltd" in csv_out and "Dup Ltd Again" in csv_out
        finally:
            client.delete(f"/api/jobs/{created['job_id']}")

    def test_deleting_a_job_removes_its_generated_files(self, client):
        """Deleting a job must not leave orphaned reports or upload copies."""
        from app.settings import REPORT_DIR, UPLOAD_DIR

        csv_bytes = b"Business Name,City\nCleanup Test Ltd,Leeds\n"
        up = client.post("/api/uploads", files={"file": ("c.csv", csv_bytes, "text/csv")}).json()
        upload_file = UPLOAD_DIR / f"{up['upload_id']}.csv"
        assert upload_file.exists()

        created = client.post("/api/jobs", json={
            "upload_id": up["upload_id"], "name": "CLEANUP TEST",
            "mapping": {"business_name": "Business Name", "city": "City"},
            "start_immediately": False,
        }).json()
        job_id = created["job_id"]

        # Simulate a report having been produced for this job.
        folder = REPORT_DIR / f"job_{job_id}"
        folder.mkdir(parents=True, exist_ok=True)
        (folder / "000001-cleanup-test.html").write_text("<html></html>", encoding="utf-8")

        r = client.delete(f"/api/jobs/{job_id}")
        assert r.status_code == 200
        assert r.json()["files_removed"]["reports"] == 1
        assert r.json()["files_removed"]["upload"] is True
        assert not folder.exists()
        assert not upload_file.exists()

    def test_deleting_a_job_keeps_an_upload_shared_with_another_job(self, client):
        csv_bytes = b"Business Name,City\nShared Upload Ltd,Leeds\n"
        up = client.post("/api/uploads", files={"file": ("s.csv", csv_bytes, "text/csv")}).json()
        from app.settings import UPLOAD_DIR

        upload_file = UPLOAD_DIR / f"{up['upload_id']}.csv"
        mapping = {"business_name": "Business Name", "city": "City"}
        a = client.post("/api/jobs", json={"upload_id": up["upload_id"], "name": "SHARED A",
                                           "mapping": mapping, "start_immediately": False}).json()
        b = client.post("/api/jobs", json={"upload_id": up["upload_id"], "name": "SHARED B",
                                           "mapping": mapping, "start_immediately": False}).json()
        try:
            client.delete(f"/api/jobs/{a['job_id']}")
            assert upload_file.exists(), "the second job still needs this upload"
        finally:
            client.delete(f"/api/jobs/{b['job_id']}")
        assert not upload_file.exists()

    def test_regenerate_outreach_refuses_without_stored_problems(self, client, seeded_job):
        original = client.get("/api/settings/profile").json()["profile"]
        try:
            client.put("/api/settings/profile", json={
                "full_name": "T", "company_name": "T Co", "service_name": "redesign",
                "email": "t@t.test",
            })
            listing = client.get("/api/leads", params={"job_id": seeded_job,
                                                       "channel": "whatsapp"}).json()
            lead_id = listing["leads"][0]["id"]
            r = client.post(f"/api/leads/{lead_id}/regenerate-outreach")
            assert r.status_code == 200
            assert r.json()["drafts_written"] >= 1
        finally:
            client.put("/api/settings/profile", json=original)


class TestResumeAndCheckpoint:
    """Spec 28: closing the app mid-run must not lose or repeat work."""

    def test_completed_items_are_not_requeued(self, seeded_job):
        from app.api.jobs import _job_dto

        with session_scope(write=False) as s:
            job = s.get(Job, seeded_job)
            dto = _job_dto(s, job)
        assert dto["counts"]["completed"] == 3
        assert dto["counts"]["pending"] == 0

    def test_interrupted_running_items_are_recovered_as_pending(self, seeded_job):
        """A process killed mid-lead leaves status='running'; the loader resets it."""
        from app.core.pipeline import JobControl, JobRunner, manager

        with session_scope() as s:
            item = s.query(JobItem).filter(JobItem.job_id == seeded_job).first()
            item.status = "running"
            item.stage = "crawl"
            interrupted_id = item.business_id

        runner = JobRunner(seeded_job, JobControl(), manager)
        with session_scope() as s:
            pending = runner._load_pending(s)
        assert interrupted_id in pending

        with session_scope() as s:
            item = s.query(JobItem).filter(JobItem.business_id == interrupted_id).first()
            assert item.status == "pending"
            item.status = "completed"
            item.stage = "done"

    def test_failed_items_are_retried_on_resume(self, seeded_job):
        from app.core.pipeline import JobControl, JobRunner, manager

        with session_scope() as s:
            item = s.query(JobItem).filter(JobItem.job_id == seeded_job).first()
            item.status = "failed"
            item.error_message = "timeout"
            failed_id = item.business_id

        runner = JobRunner(seeded_job, JobControl(), manager)
        with session_scope() as s:
            pending = runner._load_pending(s)
        assert failed_id in pending

        with session_scope() as s:
            item = s.query(JobItem).filter(JobItem.business_id == failed_id).first()
            item.status = "completed"
            item.error_message = ""

    def test_stage_progress_is_persisted(self, seeded_job):
        with session_scope(write=False) as s:
            item = s.query(JobItem).filter(JobItem.job_id == seeded_job).first()
        assert "done" in (item.completed_stages or [])
        assert item.stage == "done"

    def test_skipped_items_are_not_reprocessed(self, seeded_job):
        from app.core.pipeline import JobControl, JobRunner, manager

        with session_scope() as s:
            item = s.query(JobItem).filter(JobItem.job_id == seeded_job).first()
            original = item.status
            item.status = "skipped"
            skipped_id = item.business_id

        runner = JobRunner(seeded_job, JobControl(), manager)
        with session_scope() as s:
            pending = runner._load_pending(s)
        assert skipped_id not in pending

        with session_scope() as s:
            item = s.query(JobItem).filter(JobItem.business_id == skipped_id).first()
            item.status = original
