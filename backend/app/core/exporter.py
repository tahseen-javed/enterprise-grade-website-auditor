"""
Exports (spec 11, 33).

The final CSV reproduces every original row and every original column in the
original order, then appends the enrichment columns at the END. One input
business stays exactly one output row - enrichment never splits a lead into
multiple rows.
"""

from __future__ import annotations

import csv
import datetime as dt
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import (
    Business,
    ContactEmail,
    ContactPhone,
    Job,
    JobItem,
    OutreachDraft,
    WebsiteAudit,
)
from ..settings import EXPORT_DIR

# Appended, in this order, after every original column (spec 11).
ENRICHMENT_COLUMNS: List[str] = [
    "website_status",
    "website_final",
    "website_identity_confidence",
    "website_source",

    "email_1", "email_1_source", "email_1_status",
    "email_2", "email_2_source", "email_2_status",

    "phone_raw_original",
    "phone_normalized",
    "phone_country",
    "phone_type",
    "phone_status",

    "whatsapp_status",
    "whatsapp_reason",
    "whatsapp_url",

    "linkedin_url",
    "linkedin_status",

    "website_score",
    "opportunity_tier",
    "lead_tier",

    "problems",
    "recommendations",

    "contact_channel",
    "contact_channel_reason",

    "whatsapp_message",
    "whatsapp_draft_url",

    "email_subject",
    "email_message",
    "email_draft_url",

    "linkedin_message",

    "call_notes",

    "audit_report_path",
    "audit_status",
    "audit_error",
    "processed_at",
]

# contact_channel is stored internally lowercase (whatsapp/email/linkedin/
# phone/none), matching every other status vocabulary in this app - this is
# the export-only presentation the routing spec asks for.
_CONTACT_CHANNEL_DISPLAY: Dict[str, str] = {
    "whatsapp": "WHATSAPP",
    "email": "EMAIL",
    "linkedin": "LINKEDIN",
    "phone": "PHONE",
    "none": "SKIP",
    "": "SKIP",
    # Retained only for jobs processed before this routing existed - "contact
    # form" was a real channel value then, and relabeling it PHONE or SKIP
    # would misrepresent what was actually found for that lead.
    "website_contact": "WEBSITE_CONTACT",
}


def _fmt_dt(value: Optional[dt.datetime]) -> str:
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else ""


def _join_problems(problems: List[Dict[str, Any]]) -> str:
    return " | ".join(f"{i + 1}. {p.get('title', '')}" for i, p in enumerate(problems or []))


def _join_recommendations(recs: List[Dict[str, Any]]) -> str:
    return " | ".join(
        f"{i + 1}. {r.get('recommendation', '')}" for i, r in enumerate(recs or [])
    )


def build_rows(session, job_id: int) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Returns (headers, rows) with original columns first, enrichment last."""
    job: Optional[Job] = session.get(Job, job_id)
    if job is None:
        raise ValueError(f"Job {job_id} does not exist.")

    original_columns: List[str] = list(job.original_columns or [])

    businesses = (
        session.query(Business)
        .filter(Business.job_id == job_id)
        .order_by(Business.row_index)
        .all()
    )

    ids = [b.id for b in businesses]
    audits = {
        a.business_id: a
        for a in session.query(WebsiteAudit).filter(WebsiteAudit.business_id.in_(ids)).all()
    } if ids else {}
    phones: Dict[int, ContactPhone] = {}
    for p in (session.query(ContactPhone).filter(ContactPhone.business_id.in_(ids)).all() if ids else []):
        phones.setdefault(p.business_id, p)
    emails: Dict[int, List[ContactEmail]] = {}
    for e in (
        session.query(ContactEmail)
        .filter(ContactEmail.business_id.in_(ids))
        .order_by(ContactEmail.rank)
        .all() if ids else []
    ):
        emails.setdefault(e.business_id, []).append(e)
    drafts: Dict[int, Dict[Tuple[str, str], OutreachDraft]] = {}
    for d in (session.query(OutreachDraft).filter(OutreachDraft.business_id.in_(ids)).all() if ids else []):
        drafts.setdefault(d.business_id, {})[(d.channel, d.variant)] = d
    items = {
        i.business_id: i
        for i in (session.query(JobItem).filter(JobItem.business_id.in_(ids)).all() if ids else [])
    }

    headers = original_columns + [c for c in ENRICHMENT_COLUMNS if c not in original_columns]
    # If an original column collides with an enrichment name, disambiguate the
    # enrichment side so the user's data is never overwritten.
    collisions = {c for c in ENRICHMENT_COLUMNS if c in original_columns}
    enrich_names = {c: (f"audit_{c}" if c in collisions else c) for c in ENRICHMENT_COLUMNS}
    if collisions:
        headers = original_columns + [enrich_names[c] for c in ENRICHMENT_COLUMNS]

    rows: List[Dict[str, Any]] = []
    for b in businesses:
        row: Dict[str, Any] = {}
        raw = b.raw or {}
        for col in original_columns:
            row[col] = raw.get(col, "")

        audit = audits.get(b.id)
        phone = phones.get(b.id)
        elist = emails.get(b.id, [])
        dmap = drafts.get(b.id, {})
        item = items.get(b.id)

        wa_initial = dmap.get(("whatsapp", "initial"))
        em_initial = dmap.get(("email", "initial"))
        li_initial = dmap.get(("linkedin", "initial"))
        call_draft = dmap.get(("call", "initial"))

        e1 = elist[0] if len(elist) > 0 else None
        e2 = elist[1] if len(elist) > 1 else None

        values: Dict[str, Any] = {
            "website_status": b.website_status or "",
            "website_final": b.website_final or "",
            "website_identity_confidence": (
                "" if b.website_identity_confidence is None
                else round(b.website_identity_confidence, 3)
            ),
            "website_source": b.website_source or "",

            "email_1": e1.email if e1 else "",
            "email_1_source": e1.source_url if e1 else "",
            "email_1_status": e1.status if e1 else "",
            "email_2": e2.email if e2 else "",
            "email_2_source": e2.source_url if e2 else "",
            "email_2_status": e2.status if e2 else "",

            "phone_raw_original": (phone.phone_raw if phone else b.phone_raw) or "",
            "phone_normalized": phone.phone_normalized if phone else "",
            "phone_country": phone.phone_country if phone else "",
            "phone_type": phone.phone_type if phone else "",
            "phone_status": phone.validation_status if phone else "unavailable",

            "whatsapp_status": phone.whatsapp_status if phone else "no_phone",
            "whatsapp_reason": phone.whatsapp_reason if phone else "",
            "whatsapp_url": phone.whatsapp_url if phone else "",

            "linkedin_url": b.linkedin_url or "",
            "linkedin_status": b.linkedin_status or "not_checked",

            "website_score": "" if b.score is None else b.score,
            "opportunity_tier": b.opportunity_tier or "",
            "lead_tier": b.lead_tier or "",

            "problems": _join_problems(audit.problems if audit else []),
            "recommendations": _join_recommendations(audit.recommendations if audit else []),

            "contact_channel": _CONTACT_CHANNEL_DISPLAY.get(b.best_channel or "", (b.best_channel or "SKIP").upper()),
            "contact_channel_reason": b.channel_reason or "",

            "whatsapp_message": wa_initial.message if wa_initial else "",
            "whatsapp_draft_url": wa_initial.draft_url if wa_initial else "",

            "email_subject": em_initial.subject if em_initial else "",
            "email_message": em_initial.message if em_initial else "",
            "email_draft_url": em_initial.draft_url if em_initial else "",

            "linkedin_message": li_initial.message if li_initial else "",

            "call_notes": call_draft.message if call_draft else "",

            "audit_report_path": audit.report_path if audit else "",
            "audit_status": (
                audit.audit_status if audit
                else (item.status if item else "not_processed")
            ),
            "audit_error": (audit.audit_error if audit else (item.error_message if item else "")),
            "processed_at": _fmt_dt(b.processed_at),
        }

        for canonical, out_name in enrich_names.items():
            row[out_name] = values.get(canonical, "")
        rows.append(row)

    return headers, rows


def export_csv(session, job_id: int) -> Path:
    headers, rows = build_rows(session, job_id)
    job = session.get(Job, job_id)
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    path = EXPORT_DIR / f"job{job_id}-{_safe(job.name if job else 'export')}-{stamp}.csv"

    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({h: _cell(row.get(h, "")) for h in headers})
    return path


def export_xlsx(session, job_id: int) -> Path:
    headers, rows = build_rows(session, job_id)
    job = session.get(Job, job_id)
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    path = EXPORT_DIR / f"job{job_id}-{_safe(job.name if job else 'export')}-{stamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    original_count = len(list(job.original_columns or [])) if job else 0
    head_fill = PatternFill("solid", fgColor="1E2537")
    enrich_fill = PatternFill("solid", fgColor="3B5BDB")
    head_font = Font(color="FFFFFF", bold=True, size=10)

    ws.append(headers)
    for idx, _h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = head_fill if idx <= original_count else enrich_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center", wrap_text=False)

    for row in rows:
        ws.append([_cell(row.get(h, "")) for h in headers])

    ws.freeze_panes = "A2"
    for idx, h in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        longest = max([len(str(h))] + [len(str(r.get(h, ""))[:80]) for r in rows[:400]] or [10])
        ws.column_dimensions[letter].width = min(52, max(12, longest + 2))

    # A second sheet documenting what each appended column means.
    doc = wb.create_sheet("Enrichment key")
    doc.append(["Column", "Meaning"])
    doc["A1"].font = Font(bold=True)
    doc["B1"].font = Font(bold=True)
    for col, meaning in COLUMN_DOCS.items():
        doc.append([col, meaning])
    doc.column_dimensions["A"].width = 32
    doc.column_dimensions["B"].width = 110
    for r in range(2, doc.max_row + 1):
        doc.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)
    return path


COLUMN_DOCS: Dict[str, str] = {
    "website_status": "valid / redirected / unavailable / blocked / mismatch / not_found / "
                      "not_a_website (social or directory profile) / no_website (verified none found).",
    "website_identity_confidence": "0-1 confidence that the site belongs to this exact business, "
                                   "from domain, title, phone and address matching.",
    "website_source": "csv = supplied in your file; discovered = found by domain guess and "
                      "confirmed by identity matching; none.",
    "email_1": "A public address found on the business's own website. Never guessed or constructed.",
    "email_1_source": "The exact page URL the address was found on.",
    "email_1_status": "valid_public / mx_valid / domain_valid / syntax_valid / risky / invalid / "
                      "unknown. Mailbox deliverability is never claimed.",
    "phone_normalized": "E.164 format. Empty when the number could not be parsed with confidence.",
    "phone_status": "valid / possible / invalid / unparseable / ambiguous_region / unavailable.",
    "whatsapp_status": "confirmed_on_website (a WhatsApp link for this number is published on their "
                       "site - the only status that selects WhatsApp as the contact channel) / "
                       "usable_unverified (valid WhatsApp-capable number, informational only - a CSV "
                       "phone number is never on its own treated as WhatsApp) / unlikely (landline) / "
                       "invalid_number / no_phone.",
    "whatsapp_url": "wa.me click-to-chat link with the message pre-filled. Nothing is sent "
                    "automatically - you open and send it yourself.",
    "linkedin_url": "The business's own LinkedIn company page, found as a link on their website. "
                    "Never a guess and never a personal employee profile.",
    "linkedin_status": "not_checked (WhatsApp or email already usable, so LinkedIn was not looked "
                       "for) / found / not_found.",
    "website_score": "0-100 opportunity score. HIGHER means more measured room to improve.",
    "opportunity_tier": "Very High 90+, High 75-89, Good 60-74, Moderate 40-59, Low 0-39.",
    "lead_tier": "A+ / A / B / C / D combining opportunity, website validity, contact availability "
                 "and whether a strong specific problem exists.",
    "problems": "The detected problems, each backed by an actual measurement.",
    "contact_channel": "WHATSAPP / EMAIL / LINKEDIN / PHONE / SKIP, in that priority order. WhatsApp "
                       "requires a confirmed link on the site; email requires a public address found "
                       "on the site; LinkedIn requires the business's own company page found on the "
                       "site; phone requires a valid number; SKIP means none of the above.",
    "whatsapp_message": "Draft only. Personalised from this business's own audit findings.",
    "email_draft_url": "mailto: link with subject and body pre-filled. Not sent automatically.",
    "linkedin_message": "Draft only, for pasting into a LinkedIn message manually - there is no URL "
                        "scheme to pre-fill a LinkedIn DM.",
    "call_notes": "Opening line and talking points for leads with no WhatsApp, email or LinkedIn path.",
    "audit_status": "completed / failed / no_clear_opportunity (audited, but no meaningful problem "
                    "was found - deliberately not given an invented one).",
    "processed_at": "UTC timestamp when this row finished processing.",
}


def _cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    s = str(value)
    # Excel formula-injection guard.
    if s[:1] in ("=", "+", "-", "@", "\t", "\r"):
        s = "'" + s
    return s[:32000]


def _safe(name: str) -> str:
    import re

    return re.sub(r"[^A-Za-z0-9._-]+", "-", name or "export").strip("-")[:48] or "export"
