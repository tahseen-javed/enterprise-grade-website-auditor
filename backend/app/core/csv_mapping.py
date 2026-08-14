"""
CSV ingestion + intelligent column mapping.

Column names are never assumed. Headers are scored two ways - by name
(alias table + fuzzy match) and by the shape of the actual values in the
column - and the UI always gets the chance to correct the result.

Every original column is preserved verbatim; mapping only records *which*
original column feeds each canonical field.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import re
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from charset_normalizer import from_bytes

# --------------------------------------------------------------------------
# Canonical target fields
# --------------------------------------------------------------------------

CANONICAL_FIELDS: List[Dict[str, Any]] = [
    {"key": "business_name", "label": "Business name", "required": True},
    {"key": "category", "label": "Category / industry", "required": False},
    {"key": "phone", "label": "Phone", "required": False},
    {"key": "website", "label": "Website", "required": False},
    {"key": "google_maps_url", "label": "Google Maps URL", "required": False},
    {"key": "address", "label": "Address", "required": False},
    {"key": "city", "label": "City", "required": False},
    {"key": "state", "label": "State / region", "required": False},
    {"key": "country", "label": "Country", "required": False},
    {"key": "postal_code", "label": "Postal code", "required": False},
    {"key": "rating", "label": "Rating", "required": False},
    {"key": "review_count", "label": "Review count", "required": False},
    {"key": "place_id", "label": "Place ID", "required": False},
    {"key": "email", "label": "Email (if already present)", "required": False},
]

FIELD_KEYS = [f["key"] for f in CANONICAL_FIELDS]

ALIASES: Dict[str, List[str]] = {
    "business_name": [
        "business name", "businessname", "name", "company", "company name",
        "title", "business", "store name", "place name", "organization",
        "organisation", "listing name", "shop name", "brand",
    ],
    "category": [
        "category", "categories", "type", "business type", "industry", "niche",
        "primary category", "main category", "business category", "sector",
        "keyword", "search term", "query",
    ],
    "phone": [
        "phone", "phone number", "telephone", "tel", "mobile", "contact",
        "contact number", "phone1", "primary phone", "business phone",
        "phone_number", "cell", "whatsapp", "number",
    ],
    "website": [
        "website", "web site", "url", "site", "web", "homepage", "domain",
        "website url", "site url", "web address", "link", "business website",
    ],
    "google_maps_url": [
        "google maps url", "maps url", "google url", "gmb url", "maps link",
        "google maps link", "map url", "google_maps", "place url", "google link",
    ],
    "address": [
        "address", "full address", "street address", "street", "location",
        "address line 1", "formatted address", "addr", "complete address",
    ],
    "city": ["city", "town", "locality", "municipality", "city name"],
    "state": [
        "state", "region", "province", "county", "administrative area",
        "state province", "district",
    ],
    "country": ["country", "country name", "nation", "country code", "countrycode"],
    "postal_code": [
        "postal code", "postcode", "zip", "zip code", "zipcode", "post code",
        "postal", "pin code",
    ],
    "rating": [
        "rating", "stars", "star rating", "average rating", "score",
        "google rating", "review rating", "avg rating",
    ],
    "review_count": [
        "review count", "reviews", "number of reviews", "reviews count",
        "total reviews", "review_count", "num reviews", "ratings count",
        "user ratings total",
    ],
    "place_id": ["place id", "placeid", "google place id", "cid", "fid", "data id"],
    "email": ["email", "e-mail", "email address", "contact email", "mail"],
}

# Headers that look like a field but are the wrong one - hard blocks that stop
# "review_url" landing on "review_count" or "email_status" on "email".
NEGATIVE_HINTS: Dict[str, List[str]] = {
    "review_count": ["url", "link", "text", "date", "author"],
    "rating": ["count", "total", "number", "url"],
    "website": ["status", "score", "audit"],
    "phone": ["status", "type", "country", "code"],
    "email": ["status", "source", "valid", "verified"],
}


def _norm_header(h: str) -> str:
    h = (h or "").strip().lower()
    h = re.sub(r"[_\-./\\]+", " ", h)
    h = re.sub(r"[^a-z0-9 ]+", "", h)
    return re.sub(r"\s+", " ", h).strip()


# --------------------------------------------------------------------------
# Value-shape detectors - used when the header name is unhelpful
# --------------------------------------------------------------------------

_RE_URL = re.compile(r"^(https?://|www\.)|\.[a-z]{2,24}(/|$)", re.I)
_RE_EMAIL = re.compile(r"^[^@\s]+@[^@\s]+\.[a-z]{2,}$", re.I)
_RE_PHONE = re.compile(r"^[+()\d][\d\s()\-.+extEXT]{5,}$")
_RE_MAPS = re.compile(r"(google\.[a-z.]+/maps|goo\.gl/maps|maps\.app\.goo\.gl)", re.I)
_RE_POSTAL = re.compile(r"^[A-Z0-9][A-Z0-9 \-]{2,9}$", re.I)
_RE_PLACE_ID = re.compile(r"^(ChI|0x[0-9a-f]+:|GhIJ|EiQ)", re.I)


def _shape_scores(values: List[str]) -> Dict[str, float]:
    """Fraction of non-empty sample values matching each shape."""
    vals = [v.strip() for v in values if v and v.strip()]
    if not vals:
        return {}
    n = len(vals)

    def frac(pred) -> float:
        return sum(1 for v in vals if pred(v)) / n

    def is_float(v: str) -> bool:
        try:
            f = float(v.replace(",", "."))
            return 0.0 <= f <= 5.0
        except ValueError:
            return False

    def is_int(v: str) -> bool:
        return v.replace(",", "").replace(" ", "").isdigit()

    return {
        "google_maps_url": frac(lambda v: bool(_RE_MAPS.search(v))),
        "website": frac(lambda v: bool(_RE_URL.search(v)) and not _RE_MAPS.search(v)),
        "email": frac(lambda v: bool(_RE_EMAIL.match(v))),
        "phone": frac(lambda v: bool(_RE_PHONE.match(v)) and sum(c.isdigit() for c in v) >= 7),
        "rating": frac(is_float),
        "review_count": frac(is_int),
        "postal_code": frac(lambda v: bool(_RE_POSTAL.match(v)) and any(c.isdigit() for c in v)),
        "place_id": frac(lambda v: bool(_RE_PLACE_ID.match(v))),
    }


def _name_score(header: str, field: str) -> float:
    h = _norm_header(header)
    if not h:
        return 0.0
    aliases = ALIASES.get(field, [])

    for bad in NEGATIVE_HINTS.get(field, []):
        if bad in h and not any(h == a for a in aliases):
            return 0.0

    if h in aliases:
        return 1.0
    if h == _norm_header(field):
        return 1.0

    best = 0.0
    for alias in aliases:
        if h == alias:
            return 1.0
        # whole-token containment, e.g. "business phone number" ~ "phone"
        if alias in h or h in alias:
            tok_ratio = min(len(alias), len(h)) / max(len(alias), len(h))
            best = max(best, 0.72 + 0.18 * tok_ratio)
        else:
            ratio = SequenceMatcher(None, h, alias).ratio()
            if ratio > 0.86:
                best = max(best, ratio * 0.85)
    return min(best, 0.97)


def suggest_mapping(
    headers: List[str], sample_rows: List[Dict[str, str]]
) -> Dict[str, Any]:
    """
    Score every (header, canonical field) pair, then assign greedily so one
    header is never claimed by two fields.
    """
    columns: Dict[str, List[str]] = {
        h: [str(r.get(h, "") or "") for r in sample_rows] for h in headers
    }
    shapes = {h: _shape_scores(v) for h, v in columns.items()}

    candidates: List[Tuple[float, str, str, Dict[str, float]]] = []
    for h in headers:
        for field in FIELD_KEYS:
            ns = _name_score(h, field)
            ss = shapes.get(h, {}).get(field, 0.0)
            # Name dominates; value shape confirms or rescues an odd header.
            if ns >= 0.5 and ss >= 0.5:
                total = min(1.0, ns * 0.75 + ss * 0.35)
            elif ns >= 0.5:
                total = ns * 0.85
            elif ss >= 0.8 and field in {
                "website", "email", "phone", "google_maps_url", "place_id",
            }:
                total = ss * 0.6  # shape-only: usable but flagged low confidence
            else:
                total = 0.0
            if total > 0:
                candidates.append((total, h, field, {"name": ns, "shape": ss}))

    candidates.sort(key=lambda c: -c[0])
    mapping: Dict[str, Optional[str]] = {f: None for f in FIELD_KEYS}
    confidence: Dict[str, float] = {}
    evidence: Dict[str, Dict[str, float]] = {}
    used_headers: set = set()

    for total, header, field, parts in candidates:
        if mapping[field] is not None or header in used_headers:
            continue
        if total < 0.45:
            continue
        mapping[field] = header
        confidence[field] = round(total, 3)
        evidence[field] = {k: round(v, 3) for k, v in parts.items()}
        used_headers.add(header)

    unmapped = [h for h in headers if h not in used_headers]
    missing_required = [
        f["key"] for f in CANONICAL_FIELDS if f["required"] and not mapping.get(f["key"])
    ]
    low_conf = [f for f, c in confidence.items() if c < 0.7]

    return {
        "mapping": mapping,
        "confidence": confidence,
        "evidence": evidence,
        "unmapped_columns": unmapped,
        "missing_required": missing_required,
        "low_confidence_fields": low_conf,
        # The UI must be shown when anything is missing or shaky (spec 2).
        "needs_review": bool(missing_required or low_conf),
    }


# --------------------------------------------------------------------------
# Reading
# --------------------------------------------------------------------------


def detect_encoding(raw: bytes) -> str:
    head = raw[:200_000]
    if head.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    try:
        head.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        pass
    try:
        best = from_bytes(head).best()
        if best and best.encoding:
            return best.encoding
    except Exception:
        pass
    return "latin-1"


def detect_dialect(text_sample: str) -> csv.Dialect:
    try:
        return csv.Sniffer().sniff(text_sample, delimiters=",;\t|")
    except csv.Error:
        class _D(csv.Dialect):
            delimiter = ","
            quotechar = '"'
            doublequote = True
            skipinitialspace = False
            lineterminator = "\r\n"
            quoting = csv.QUOTE_MINIMAL

        return _D()


class CsvReadError(Exception):
    pass


# --------------------------------------------------------------------------
# Excel ingestion - converted to CSV text once, at the door, so every
# downstream consumer (preview, remap, job creation) only ever has to know
# one format. No mapping/scoring logic is duplicated for Excel.
# --------------------------------------------------------------------------

EXCEL_SUFFIXES = (".xlsx", ".xlsm")


def _excel_cell_to_str(value: Any) -> str:
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    return str(value)


def excel_to_csv_text(raw: bytes) -> str:
    """Read the first worksheet of an .xlsx/.xlsm workbook and return it as CSV text."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise CsvReadError(f"The Excel file could not be opened: {exc}") from None

    try:
        ws = wb.active
        if ws is None:
            raise CsvReadError("The Excel file has no worksheets.")

        buf = io.StringIO(newline="")
        writer = csv.writer(buf)
        wrote_any = False
        for row in ws.iter_rows(values_only=True):
            if row is None or all(v is None for v in row):
                continue
            writer.writerow(["" if v is None else _excel_cell_to_str(v) for v in row])
            wrote_any = True
        if not wrote_any:
            raise CsvReadError("The Excel sheet has no data rows.")
        return buf.getvalue()
    finally:
        wb.close()


def read_csv(path: Path, limit: Optional[int] = None) -> Tuple[List[str], List[Dict[str, str]]]:
    """Read a CSV into (headers, rows). Rows keep every original column."""
    raw = path.read_bytes()
    if not raw.strip():
        raise CsvReadError("The file is empty.")

    encoding = detect_encoding(raw)
    try:
        text_data = raw.decode(encoding, errors="replace")
    except LookupError:
        text_data = raw.decode("utf-8", errors="replace")

    dialect = detect_dialect(text_data[:8192])
    reader = csv.reader(io.StringIO(text_data, newline=""), dialect)

    try:
        header_row = next(reader)
    except StopIteration:
        raise CsvReadError("The file has no header row.") from None

    headers: List[str] = []
    seen: Dict[str, int] = {}
    for i, h in enumerate(header_row):
        name = (h or "").strip().lstrip("﻿") or f"column_{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)

    rows: List[Dict[str, str]] = []
    width = len(headers)
    for values in reader:
        if not any((v or "").strip() for v in values):
            continue
        if len(values) < width:
            values = list(values) + [""] * (width - len(values))
        row = {headers[i]: (values[i] if i < len(values) else "") for i in range(width)}
        # Preserve overflow columns rather than dropping data.
        if len(values) > width:
            row["_extra_columns"] = " | ".join(str(v) for v in values[width:])
        rows.append(row)
        if limit and len(rows) >= limit:
            break

    if not rows:
        raise CsvReadError("The file contains a header but no data rows.")
    if any("_extra_columns" in r for r in rows) and "_extra_columns" not in headers:
        headers.append("_extra_columns")

    return headers, rows


def preview_csv(path: Path, sample: int = 25) -> Dict[str, Any]:
    headers, rows = read_csv(path, limit=sample)
    total = count_rows(path)
    suggestion = suggest_mapping(headers, rows)
    return {
        "headers": headers,
        "sample_rows": rows[:sample],
        "row_count": total,
        **suggestion,
    }


def count_rows(path: Path) -> int:
    raw = path.read_bytes()
    encoding = detect_encoding(raw)
    text_data = raw.decode(encoding, errors="replace")
    dialect = detect_dialect(text_data[:8192])
    reader = csv.reader(io.StringIO(text_data, newline=""), dialect)
    n = -1  # discount header
    for values in reader:
        if any((v or "").strip() for v in values):
            n += 1
    return max(0, n)
